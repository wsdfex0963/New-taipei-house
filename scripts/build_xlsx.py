#!/usr/bin/env python3
"""
產生「新北市建案 看屋檢查清單」比較表 Excel。

輸入 JSON 結構 (--data)：
{
  "search_date": "2026-08-24",
  "sources": "591新建案、591實價登錄、樂居、住展、5168比價王...",
  "projects": [
    {
      "name": "民生新埔",
      "address": "新北市板橋區民生路三段186號附近",
      "metro": "捷運新埔站(步行5分鐘)",
      "data": {
        "建案名稱": "民生新埔",
        "權狀坪數": "套房12坪、1+1房15坪\n**2房18~28坪、3房30~32坪**",
        "...": "..."
      }
    }
  ]
}

標記規則：
  - `**xxx**` 包住的文字 -> 藍字粗體（用於 > 2 房的資訊）
  - 字串等於「未知待查詢」或以此開頭 -> 整格紅字
  - `【現場確認】` -> 紅字粗體（公開資料查不到，須到接待中心問或調謄本）
  - `【試算】`     -> 橘字粗體（依公式推算，非官方公告值）
  - `⚠️` 起至行尾  -> 紅字（風險提醒）
  - `✅` 起至行尾  -> 綠字（正面條件）
  - 其餘 -> 一般黑字

用法：
  python3 scripts/build_xlsx.py --data projects_data.json --output out.xlsx
"""
import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

TITLE = "新北市建案 看屋檢查清單"

CATEGORIES = [
    ("一、基本資料", [
        "建案名稱", "建案地址", "建照/使照字號", "建商", "代銷",
        "基地面積", "總戶數／棟數／樓層規劃", "產品規劃（房型／格局）",
    ]),
    ("二、建商與代銷背景", [
        "建商過往作品", "建商財務／信用評價", "建商是否有糾紛或訴訟紀錄",
        "代銷公司背景", "保固條款（結構／防水年限）", "履約保證方式",
        "客戶評價／網路口碑",
    ]),
    ("三、產品規劃與坪數", [
        "各房型權狀坪數區間", "主建物／附屬建物／公設比例", "公設項目與內容",
        "車位規劃（平面／機械／坡道）與車位價格", "樓層規劃（地上／地下層數）",
        "建材設備（廚具、衛浴、電梯品牌）", "格局方正度／採光通風",
        "是否有夾層或樓中樓", "交屋標準與客變條件",
    ]),
    ("四、生活機能與交通", [
        "鄰近捷運站與步行時間", "公車路線與班次", "鄰近學區（國小／國中）",
        "鄰近市場／超市／賣場", "鄰近公園綠地", "鄰近醫院診所",
        "未來重大建設／都更計畫", "生活機能綜合評分",
    ]),
    ("五、嫌惡設施與環境風險", [
        "鄰近嫌惡設施（殯儀館、垃圾場、高壓電塔等）", "淹水潛勢", "土壤液化潛勢",
        "是否位於斷層帶", "噪音來源（鐵路／快速道路／工廠）", "治安狀況",
        "實價登錄行情（周邊成交價）", "未來供給量（周邊新建案數）",
    ]),
    ("六、財務評估", [
        "總價區間", "單價區間（每坪）", "貸款成數與利率試算",
        "頭期款與付款方式（工程期款）", "稅費估算（契稅／印花稅／代書費）",
        "管理費（每坪／月）", "增值潛力／轉手性評估",
    ]),
]

UNKNOWN_MARK = "未知待查詢"

DARK_BLUE = "1F3864"
LIGHT_BLUE = "BDD7EE"
RED = "C00000"
BLUE_BOLD = "1F4E78"
CALC_ORANGE = "BF8F00"
GREEN = "1E7145"

thin = Side(style="thin", color="B7B7B7")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


TOKEN_RE = re.compile(
    r"(\*\*[^*]+\*\*"          # **超出看屋範圍**
    r"|【現場確認】"            # 須到接待中心／調謄本才問得到
    r"|【試算】"                # 依公式推算，非官方公告值
    r"|⚠️[^\n]*"                # 風險提醒，整行標色
    r"|✅[^\n]*)"               # 正面條件，整行標色
)


def colored_runs(cell, text):
    """依 **xxx**／【現場確認】／【試算】／⚠️／✅／未知待查詢 規則拆成 rich-text 片段上色。"""
    if text is None:
        text = ""
    text = str(text)

    if text == UNKNOWN_MARK or text.startswith(UNKNOWN_MARK):
        cell.value = text
        cell.font = Font(color=RED, size=10)
        return

    parts = [p for p in TOKEN_RE.split(text) if p]
    if len(parts) > 1 or (parts and TOKEN_RE.fullmatch(parts[0])):
        blocks = []
        for p in parts:
            if p.startswith("**") and p.endswith("**"):
                f = InlineFont(color=BLUE_BOLD, b=True, sz=10)
                p = p[2:-2]
            elif p == "【現場確認】":
                f = InlineFont(color=RED, b=True, sz=10)
            elif p == "【試算】":
                f = InlineFont(color=CALC_ORANGE, b=True, sz=10)
            elif p.startswith("⚠️"):
                f = InlineFont(color=RED, sz=10)
            elif p.startswith("✅"):
                f = InlineFont(color=GREEN, sz=10)
            else:
                f = InlineFont(color="FF000000", sz=10)
            blocks.append(TextBlock(f, p))
        cell.value = CellRichText(blocks)
        cell.font = Font(color="000000", size=10)
        return

    cell.value = text
    cell.font = Font(color="000000", size=10)


def build_workbook(payload, categories=None, items=None):
    """categories 為 1-based 分類編號清單，None 表示全部六大分類。

    items 為 (start, end) 的 1-based 閉區間，只在單一分類時使用，
    用來把項目太多、檔案太大的分類再切成幾份（見 SKILL.md 的上傳上限）。
    """
    projects = payload.get("projects", [])
    search_date = payload.get("search_date", "")
    sources = payload.get("sources", "")

    selected = CATEGORIES if categories is None else [
        CATEGORIES[i - 1] for i in categories
    ]
    if items:
        if len(selected) != 1:
            raise SystemExit("--items 只能搭配單一 --categories 使用")
        name, all_items = selected[0]
        start, end = items
        selected = [(f"{name}（第{start}~{end}項）", all_items[start - 1:end])]

    n_cols = 2 + len(projects)  # A=分類 B=項目 C.. = 各建案

    wb = Workbook()
    ws = wb.active
    ws.title = "看屋檢查清單"

    # 第1列：標題
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    title_cell = ws.cell(row=1, column=1, value=TITLE)
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor=DARK_BLUE)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # 第2列：搜尋日期 + 來源
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    info_cell = ws.cell(row=2, column=1, value=f"資料搜尋日期：{search_date}　|　來源：{sources}")
    info_cell.font = Font(size=9, italic=True, color="595959")
    info_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    # 第3列：表頭
    ws.cell(row=3, column=1, value="分類").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=3, column=2, value="檢查項目").font = Font(bold=True, color="FFFFFF")
    for col in (1, 2):
        c = ws.cell(row=3, column=col)
        c.fill = PatternFill("solid", fgColor=DARK_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER

    for i, proj in enumerate(projects):
        col = 3 + i
        header = f"{proj.get('name', '')}\n{proj.get('address', '')}\n{proj.get('metro', '')}"
        c = ws.cell(row=3, column=col, value=header)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=DARK_BLUE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[3].height = 55

    # 資料列
    row = 4
    for cat_name, items in selected:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
        cat_cell = ws.cell(row=row, column=1, value=cat_name)
        cat_cell.font = Font(bold=True, size=11)
        cat_cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        cat_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20
        row += 1

        for item in items:
            ws.cell(row=row, column=1).border = BORDER
            item_cell = ws.cell(row=row, column=2, value=item)
            item_cell.font = Font(size=10, bold=True)
            item_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            item_cell.border = BORDER

            for i, proj in enumerate(projects):
                col = 3 + i
                value = proj.get("data", {}).get(item, UNKNOWN_MARK)
                cell = ws.cell(row=row, column=col)
                colored_runs(cell, value)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.border = BORDER

            ws.row_dimensions[row].height = 45
            row += 1

    # 欄寬
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 18
    for i in range(len(projects)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 38

    ws.freeze_panes = "C4"
    return wb


def slim_xlsx(path):
    """移除 xl/theme/theme1.xml 讓檔案縮小約 18%，供 Drive 內嵌上傳使用。

    openpyxl 產生的 styles.xml 只有預設字型參照 theme（<color theme="1"/>），
    先換成實色再移除 theme 部件，避免留下懸空參照讓 Excel 判定檔案損毀。
    """
    import zipfile

    src = Path(path)
    tmp = src.with_suffix(".slim.tmp")
    with zipfile.ZipFile(src) as zin, zipfile.ZipFile(
        tmp, "w", zipfile.ZIP_DEFLATED
    ) as zout:
        for item in zin.infolist():
            if item.filename == "xl/theme/theme1.xml":
                continue
            data = zin.read(item.filename)
            if item.filename == "xl/styles.xml":
                data = data.replace(b'<color theme="1" />', b'<color rgb="FF000000" />')
                data = data.replace(b'<color theme="1"/>', b'<color rgb="FF000000"/>')
            elif item.filename == "[Content_Types].xml":
                data = re.sub(
                    rb'<Override PartName="/xl/theme/theme1\.xml"[^>]*/>', b"", data
                )
            elif item.filename == "xl/_rels/workbook.xml.rels":
                data = re.sub(rb'<Relationship[^>]*theme1\.xml"[^>]*/>', b"", data)
            zout.writestr(item, data)
    tmp.replace(src)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--data", required=True, help="輸入 JSON 檔路徑")
    ap.add_argument("--output", required=True, help="輸出 xlsx 檔路徑")
    ap.add_argument(
        "--categories",
        help="只輸出指定的分類（1-based，逗號分隔，例如 1,2,3）。"
        "用於把報表拆成多份較小的檔案；省略則輸出全部六大分類。",
    )
    ap.add_argument(
        "--items",
        help="只輸出該分類的第 N~M 項（1-based 閉區間，例如 1-5）。"
        "僅能搭配單一 --categories，用於項目太多、檔案超過上傳上限時再往下拆。",
    )
    ap.add_argument(
        "--slim",
        action="store_true",
        help="移除 theme 部件讓檔案縮小約18%%，供 Google Drive 內嵌上傳（不影響顯示）。",
    )
    args = ap.parse_args()

    payload = json.loads(Path(args.data).read_text(encoding="utf-8"))
    cats = None
    if args.categories:
        cats = [int(x) for x in args.categories.split(",") if x.strip()]
    rng = None
    if args.items:
        a, b = args.items.split("-")
        rng = (int(a), int(b))
    wb = build_workbook(payload, categories=cats, items=rng)

    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    if args.slim:
        slim_xlsx(out_path)
    print(str(out_path))


if __name__ == "__main__":
    sys.exit(main())
